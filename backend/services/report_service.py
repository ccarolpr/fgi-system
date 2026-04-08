"""
Geração do relatório Excel do FGI.

O arquivo gerado é idêntico ao modelo '2026.3 - FG INCERTO CEINT.xlsx':
  - Mesma estrutura de colunas A–X
  - Mesma formatação (moeda BRL, datas, horas)
  - Mesma linha de título e cabeçalho
  - Mesma linha de total ao final

Regra absoluta: NENHUM campo extra, alerta ou flag no arquivo de saída.
Alertas existem apenas na tela de prévia (preview_service).
"""

import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.services.preview_service import LinhaPrevia, PreviaResult
from backend.config import ENCARGOS, IMPOSTOS

# Nomes dos meses em português (índice 1-12)
MESES_PT = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
    5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO",
}

# Formato moeda BRL
FMT_MOEDA = '#,##0.00'
FMT_PCT   = '0.000'
FMT_DATA  = 'DD/MM/YYYY'

# Cores (replicando visual da planilha)
COR_HEADER_DARK  = "1F3864"   # azul escuro — título
COR_HEADER_MED   = "2E75B6"   # azul médio — cabeçalho de colunas
COR_TOTAL_BG     = "D6E4F0"   # azul claro — linha de total
COR_BRANCO       = "FFFFFF"
COR_TEXTO_CLARO  = "FFFFFF"


def _borda_fina():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _celula(ws, row: int, col: int, valor, fmt: Optional[str] = None,
            bold: bool = False, bg: Optional[str] = None,
            cor_texto: Optional[str] = None, alinhamento: str = "center"):
    cell = ws.cell(row=row, column=col, value=valor)
    cell.font = Font(
        bold=bold,
        color=cor_texto or "000000",
        name="Calibri",
        size=10,
    )
    cell.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=True)
    cell.border = _borda_fina()
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    return cell


def gerar_excel(previa: PreviaResult, ano: int, mes: int, cliente: str = "CEINT") -> bytes:
    """
    Gera o arquivo Excel do FGI a partir da prévia consolidada.

    Args:
        previa:   Resultado do preview_service.gerar_previa().
        ano:      Ano do período.
        mes:      Mês do período (1–12).
        cliente:  Nome do cliente (ex: "CEINT").

    Returns:
        Bytes do arquivo .xlsx pronto para download.
    """
    mes_nome = MESES_PT.get(mes, str(mes))
    ano_curto = str(ano)[2:]  # "2026" → "26"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"F.G {mes:02d}.{ano_curto} - INC {cliente} {mes_nome[:3]}{ano_curto}"

    # -----------------------------------------------------------------------
    # Linha 1: título
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:X1")
    titulo = ws["A1"]
    titulo.value = f"FATO GERADOR INCERTO   |  {mes_nome}/{ano} - {cliente}"
    titulo.font = Font(bold=True, color=COR_TEXTO_CLARO, name="Calibri", size=12)
    titulo.fill = PatternFill(fill_type="solid", fgColor=COR_HEADER_DARK)
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24

    # -----------------------------------------------------------------------
    # Linha 2: cabeçalhos de colunas (A–X)
    # -----------------------------------------------------------------------
    CABECALHOS = [
        "",                                          # A  (vazio)
        "NOME",                                      # B
        "FUNÇÃO",                                    # C
        "CPF",                                       # D
        "ADMISSÃO",                                  # E
        "DEMISSÃO",                                  # F
        "VALOR HORA",                                # G
        "VALOR DIA",                                 # H
        "SALARIO PROJETADO",                         # I
        "QTDE DIAS ATESTADOS MÊS ANTERIOR",         # J
        "QTDE DIAS ATESTADOS 2 MESES ATRÁS",        # K
        "QTDE ATESTADOS RESTANTE DO ANO",            # L
        "QUANTIDADE DE ATESTADOS",                   # M
        "HORAS DECLARAÇÃO",                          # N
        "HORAS DECLARAÇÃO (DECIMAL)",                # O
        "VALOR TOTAL DE DECLARAÇÃO",                 # P
        "VALOR TOTAL DE ATESTADOS",                  # Q
        "TOTAL FATO GERADOR",                        # R
        "ENCARGOS",                                  # S
        "CUSTOS INDIRETOS",                          # T
        "LUCRO",                                     # U
        "TOTAL S/ IMPOSTO",                          # V
        "IMPOSTOS",                                  # W
        "TOTAL",                                     # X
    ]

    ws.row_dimensions[2].height = 40
    for col_idx, cab in enumerate(CABECALHOS, start=1):
        _celula(ws, 2, col_idx, cab, bold=True, bg=COR_HEADER_MED, cor_texto=COR_TEXTO_CLARO)

    # -----------------------------------------------------------------------
    # Linhas de dados (uma por colaborador)
    # -----------------------------------------------------------------------
    ROW_INICIO = 3

    for i, linha in enumerate(previa.linhas):
        row = ROW_INICIO + i

        def c(col: int, valor, fmt=None):
            _celula(ws, row, col, valor, fmt=fmt, alinhamento="center")

        c(1, "")                                               # A
        _celula(ws, row, 2, linha.nome, alinhamento="left")   # B NOME
        c(3, linha.funcao)                                     # C FUNÇÃO
        c(4, linha.cpf)                                        # D CPF

        # E ADMISSÃO
        _celula(ws, row, 5, linha.data_admissao, fmt=FMT_DATA)
        # F DEMISSÃO
        _celula(ws, row, 6, linha.data_demissao, fmt=FMT_DATA)

        c(7,  linha.valor_hora,               FMT_MOEDA)       # G VALOR HORA
        c(8,  linha.valor_dia,                FMT_MOEDA)       # H VALOR DIA
        c(9,  linha.salario,                  FMT_MOEDA)       # I SALARIO PROJETADO

        # J, K, L — histórico de outros meses (não disponível no novo sistema)
        c(10, None)
        c(11, None)
        c(12, None)

        c(13, linha.dias_atestado if linha.dias_atestado else None)  # M QTDE ATESTADOS

        # N HORAS DECLARAÇÃO — exibe no formato HH:MM se disponível
        if linha.horas_declaradas_raw:
            partes = linha.horas_declaradas_raw.split(":")
            if len(partes) == 2:
                from datetime import time as dtime
                try:
                    t = dtime(int(partes[0]), int(partes[1]))
                    ws.cell(row=row, column=14).value = t
                    ws.cell(row=row, column=14).number_format = "h:mm"
                    ws.cell(row=row, column=14).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=row, column=14).border = _borda_fina()
                except (ValueError, OverflowError):
                    c(14, linha.horas_declaradas_raw)
            else:
                c(14, linha.horas_declaradas_raw)
        else:
            c(14, None)

        c(15, linha.horas_declaradas_decimal if linha.horas_declaradas_decimal else None)  # O DECIMAL
        c(16, linha.custo_horas,  FMT_MOEDA)  # P VALOR TOTAL DECLARAÇÃO
        c(17, linha.custo_dias,   FMT_MOEDA)  # Q VALOR TOTAL ATESTADOS
        c(18, linha.total_fgi,    FMT_MOEDA)  # R TOTAL FATO GERADOR
        c(19, ENCARGOS,           FMT_PCT)    # S ENCARGOS (constante)
        c(20, 0,                  FMT_PCT)    # T CUSTOS INDIRETOS
        c(21, 0,                  FMT_PCT)    # U LUCRO
        c(22, linha.com_encargos, FMT_MOEDA)  # V TOTAL S/ IMPOSTO
        c(23, IMPOSTOS,           FMT_PCT)    # W IMPOSTOS (constante)
        c(24, linha.total_final,  FMT_MOEDA)  # X TOTAL

    # -----------------------------------------------------------------------
    # Linha de TOTAL
    # -----------------------------------------------------------------------
    row_total = ROW_INICIO + len(previa.linhas)
    ws.row_dimensions[row_total].height = 18

    for col in range(1, 25):
        ws.cell(row=row_total, column=col).fill = PatternFill(fill_type="solid", fgColor=COR_TOTAL_BG)
        ws.cell(row=row_total, column=col).border = _borda_fina()

    if previa.totais:
        _celula(ws, row_total, 18, previa.totais.total_fgi,          FMT_MOEDA, bold=True, bg=COR_TOTAL_BG)
        _celula(ws, row_total, 22, previa.totais.total_com_encargos, FMT_MOEDA, bold=True, bg=COR_TOTAL_BG)
        _celula(ws, row_total, 24, previa.totais.total_final,        FMT_MOEDA, bold=True, bg=COR_TOTAL_BG)

    # -----------------------------------------------------------------------
    # Largura das colunas
    # -----------------------------------------------------------------------
    LARGURAS = {
        1: 4,   # A
        2: 38,  # B NOME
        3: 30,  # C FUNÇÃO
        4: 16,  # D CPF
        5: 12,  # E ADMISSÃO
        6: 12,  # F DEMISSÃO
        7: 12,  # G VALOR HORA
        8: 12,  # H VALOR DIA
        9: 16,  # I SALARIO
        10: 8, 11: 8, 12: 8,  # J K L
        13: 10,  # M QTDE
        14: 10,  # N HORAS
        15: 10,  # O DECIMAL
        16: 14,  # P
        17: 14,  # Q
        18: 14,  # R TOTAL FGI
        19: 10,  # S
        20: 10,  # T
        21: 10,  # U
        22: 14,  # V
        23: 10,  # W
        24: 14,  # X TOTAL
    }
    for col, largura in LARGURAS.items():
        ws.column_dimensions[get_column_letter(col)].width = largura

    # Congela painel: mantém título + cabeçalho fixos ao rolar
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
