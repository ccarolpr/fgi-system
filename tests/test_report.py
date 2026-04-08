"""
Testes da prévia e geração de Excel.

Valida:
  - preview_service consolida e calcula corretamente
  - totais da prévia batem com a planilha de março/2026
  - Excel gerado contém as colunas e valores corretos
  - período segue o fluxo de status correto
  - alertas aparecem na prévia mas não no Excel
"""

import sys
import os
import io
import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.database import Base
from backend.models.employee import Employee, StatusColaborador
from backend.models.atestado import Atestado, StatusAtestado, OrigemAtestado
from backend.models.period import Period, StatusPeriodo
from backend.models.employee_audit import EmployeeAudit, OrigemAlteracao
from backend.services.preview_service import gerar_previa
from backend.services.report_service import gerar_excel


# ---------------------------------------------------------------------------
# Fixture: banco em memória com dados de março/2026
# ---------------------------------------------------------------------------

DADOS_MAR26 = [
    ("ANA CLARA DE MORAES",            "AUXILIAR OPERACIONAL I PLENO",  "49855600819", 2150, 1,  None),
    ("AURA CRISTINA VALLES RODRIGUEZ", "AUXILIAR OPERACIONAL I JUNIOR", "07215983900", 2050, 0,  "01:50"),
    ("BIANCA LIMA DA SILVA",           "AUXILIAR OPERACIONAL I JUNIOR", "54397984808", 2050, 1,  "05:13"),
    ("CARLA VIVIANE VICENTE",          "AUXILIAR OPERACIONAL I JUNIOR", "26380818847", 2050, 0,  "03:00"),
    ("CELIA PEREIRA KNUPFER RODRIGUES","AUXILIAR OPERACIONAL I PLENO",  "36918934813", 2150, 0,  "04:00"),
    ("ELEDIANE FATIMA DA SILVA SOUZA", "AUXILIAR OPERACIONAL I JUNIOR", "31942416806", 2050, 0,  "03:30"),
    ("FABIO HENRIQUE COSTA",           "AUXILIAR OPERACIONAL I JUNIOR", "10984086676", 2050, 3,  None),
    ("FELIPE AREDES SANTOS",           "AUXILIAR OPERACIONAL I JUNIOR", "35018938846", 2050, 2,  None),
    ("FERNANDO FARIA LIMA",            "AUXILIAR OPERACIONAL I JUNIOR", "44400763862", 2050, 12, None),
    ("FERNANDO PEREIRA DA SILVA",      "AUXILIAR OPERACIONAL I JUNIOR", "41238758827", 2050, 2,  None),
    ("GABRIEL LUCAS MENDES COTRIN",    "AUXILIAR OPERACIONAL I JUNIOR", "40760666857", 2050, 0,  "01:30"),
    ("GABRIEL PRADO JUSTO",            "AUXILIAR OPERACIONAL I JUNIOR", "47416381894", 2050, 1,  None),
    ("JAMILLE DE SOUZA FERREIRA",      "AUXILIAR OPERACIONAL I JUNIOR", "50302224807", 2050, 6,  "01:00"),
    ("JEFERSON VIEIRA DE FRANCA",      "AUXILIAR OPERACIONAL I JUNIOR", "23591936863", 2050, 0,  "00:52"),
    ("KLEIBER DO NASCIMENTO LUCAS",    "AUXILIAR OPERACIONAL I JUNIOR", "03411714417", 2050, 2,  None),
    ("LETICIA PEREIRA DE ALMEIDA",     "AUXILIAR OPERACIONAL I JUNIOR", "37457544844", 2050, 0,  "04:00"),
    ("LUANA FERREIRA DE SOUZA",        "AUXILIAR OPERACIONAL I JUNIOR", "41526409879", 2050, 5,  None),
    ("MARIA LEONICE DE ANDRADE SILVA", "AUXILIAR OPERACIONAL I PLENO",  "94895520625", 2150, 2,  None),
    ("MARISA SOUZA SILVA",             "AUXILIAR OPERACIONAL I JUNIOR", "86115401585", 2050, 0,  "01:53"),
    ("MAYARA APARECIDA SANTOS",        "AUXILIAR OPERACIONAL I JUNIOR", "23057721818", 2050, 1,  None),
    ("MICHELLY VICENTE SOUZA",         "AUXILIAR OPERACIONAL I JUNIOR", "32873456809", 2050, 15, None),
    ("PRISCILA NICODEMOS DA SILVA",    "SUPERVISOR OPERACIONAL JUNIOR", "11016911661", 2500, 10, "02:00"),
    ("RAQUEL DA SILVA LIMA",           "AUXILIAR OPERACIONAL I PLENO",  "37491703801", 2150, 7,  "02:20"),
    ("RENATA NAIR DE CARVALHO SILVA",  "AUXILIAR OPERACIONAL I JUNIOR", "15177910652", 2050, 3,  None),
    ("ROSEANE DA SILVA LIMA",          "AUXILIAR OPERACIONAL I JUNIOR", "39496768806", 2050, 0,  "02:36"),
    ("ROSELI GOMES DA SILVA",          "AUXILIAR OPERACIONAL I PLENO",  "57565635987", 2150, 2,  None),
    ("RUBEM FERREIRA DA CRUZ",         "AUXILIAR OPERACIONAL I PLENO",  "57986932520", 2150, 0,  "04:00"),
    ("SILVANA DOS REIS SILVA",         "AUXILIAR OPERACIONAL I PLENO",  "06627540509", 2150, 0,  "02:35"),
    ("TAMILES DE JESUS SANTOS",        "AUXILIAR OPERACIONAL I JUNIOR", "08212844528", 2050, 6,  "04:58"),
    ("TAMIRES OLIVEIRA ROCHA",         "AUXILIAR OPERACIONAL I JUNIOR", "13869591625", 2050, 1,  None),
    ("THAIS SOARES GOMES",             "AUXILIAR OPERACIONAL I JUNIOR", "12482598664", 2050, 1,  "05:17"),
    ("THAMIRES KAWANY EGYDIO FERREIRA","AUXILIAR OPERACIONAL I JUNIOR", "49807568838", 2050, 2,  None),
    ("VIVIANE MARIA DA SILVA",         "AUXILIAR OPERACIONAL I JUNIOR", "33260353810", 2050, 3,  "01:30"),
    ("WALKER SOUZA FERNANDES CAMPOS",  "AUXILIAR OPERACIONAL I JUNIOR", "46017032801", 2050, 7,  None),
]


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    from backend.models import employee, employee_audit, atestado, atestado_correcao, period  # noqa
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def db_com_dados(db):
    """Banco populado com os 34 colaboradores e atestados de março/2026."""
    for nome, funcao, cpf, salario, dias, horas in DADOS_MAR26:
        emp = Employee(cpf=cpf, nome=nome, funcao=funcao, salario=salario,
                       status=StatusColaborador.ativo)
        db.add(emp)
        db.flush()

        from backend.services.hours_parser import parse_horas
        horas_dec = parse_horas(horas) if horas else 0.0

        a = Atestado(
            colaborador_id=emp.id,
            data_inicio=datetime.date(2026, 3, 5),
            dias=dias,
            horas_declaradas_raw=horas,
            horas_declaradas_decimal=horas_dec,
            status=StatusAtestado.confirmado,
            origem=OrigemAtestado.manual,
            confirmado_por="teste",
            confirmado_em=datetime.datetime(2026, 3, 10),
        )
        db.add(a)

    db.commit()
    return db


# ---------------------------------------------------------------------------
# Testes da prévia
# ---------------------------------------------------------------------------

def test_previa_retorna_34_colaboradores(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    assert len(previa.linhas) == 34


def test_previa_total_fgi_correto(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    assert previa.totais is not None
    assert abs(previa.totais.total_fgi - 7173.76) < 0.05


def test_previa_total_com_encargos_correto(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    assert abs(previa.totais.total_com_encargos - 11635.84) < 0.05


def test_previa_total_final_correto(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    assert abs(previa.totais.total_final - 12161.78) < 0.05


def test_previa_pode_gerar_sem_pendencias(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    assert previa.pode_gerar is True
    assert previa.pendencias == []


def test_previa_bloqueia_com_pendentes(db, db_com_dados):
    # Adiciona atestado pendente ao período
    emp = db_com_dados.query(Employee).first()
    a = Atestado(
        colaborador_id=emp.id,
        data_inicio=datetime.date(2026, 3, 20),
        dias=1,
        horas_declaradas_decimal=0.0,
        status=StatusAtestado.pendente_validacao,
        origem=OrigemAtestado.manual,
    )
    db_com_dados.add(a)
    db_com_dados.commit()

    previa = gerar_previa(2026, 3, db_com_dados)
    assert previa.pode_gerar is False
    assert len(previa.pendencias) > 0


def test_previa_alerta_volume_alto(db_com_dados):
    """Michelly Vicente tem 15 dias → deve gerar alerta de volume alto."""
    previa = gerar_previa(2026, 3, db_com_dados)
    alertas_volume = [a for a in previa.alertas if a.tipo == "volume_alto"]
    assert any("MICHELLY" in a.colaborador_nome.upper() for a in alertas_volume)


def test_previa_alerta_salario_alterado(db_com_dados):
    """Colaborador com alerta_ativo=True no audit deve aparecer nos alertas da prévia."""
    emp = db_com_dados.query(Employee).filter_by(cpf="54397984808").first()
    audit = EmployeeAudit(
        employee_id=emp.id,
        campo="salario",
        valor_anterior="2050.0",
        valor_novo="3000.0",
        origem=OrigemAlteracao.importacao_planilha,
        alerta_ativo=True,
    )
    db_com_dados.add(audit)
    db_com_dados.commit()

    previa = gerar_previa(2026, 3, db_com_dados)
    alertas_salario = [a for a in previa.alertas if a.tipo == "salario_alterado"]
    assert any("54397984808" in a.colaborador_cpf for a in alertas_salario)


def test_colaboradores_sem_atestado_nao_aparecem(db_com_dados):
    """Colaborador sem atestado confirmado no período não deve estar na prévia."""
    emp_extra = Employee(
        cpf="99999999999", nome="SEM ATESTADO", funcao="AUX JR",
        salario=2050, status=StatusColaborador.ativo
    )
    db_com_dados.add(emp_extra)
    db_com_dados.commit()

    previa = gerar_previa(2026, 3, db_com_dados)
    nomes = [l.nome for l in previa.linhas]
    assert "SEM ATESTADO" not in nomes
    assert len(previa.linhas) == 34  # continua 34


# ---------------------------------------------------------------------------
# Testes do Excel gerado
# ---------------------------------------------------------------------------

def test_excel_gerado_tem_estrutura_correta(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    xlsx_bytes = gerar_excel(previa, 2026, 3)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Linha 1: título
    titulo = ws.cell(1, 1).value
    assert titulo is not None
    assert "FATO GERADOR INCERTO" in str(titulo)
    assert "CEINT" in str(titulo)

    # Linha 2: cabeçalho — verifica algumas colunas chave
    assert ws.cell(2, 2).value == "NOME"
    assert ws.cell(2, 13).value == "QUANTIDADE DE ATESTADOS"
    assert ws.cell(2, 18).value == "TOTAL FATO GERADOR"
    assert ws.cell(2, 24).value == "TOTAL"


def test_excel_total_correto_na_linha_de_total(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    xlsx_bytes = gerar_excel(previa, 2026, 3)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    row_total = 3 + len(previa.linhas)
    total_x = ws.cell(row_total, 24).value

    assert total_x is not None
    assert abs(float(total_x) - 12161.78) < 0.05


def test_excel_nao_contem_alertas(db_com_dados):
    """O Excel não deve conter campos de alerta — apenas dados de negócio."""
    emp = db_com_dados.query(Employee).filter_by(cpf="54397984808").first()
    audit = EmployeeAudit(
        employee_id=emp.id, campo="salario",
        valor_anterior="2050.0", valor_novo="3000.0",
        origem=OrigemAlteracao.importacao_planilha, alerta_ativo=True,
    )
    db_com_dados.add(audit)
    db_com_dados.commit()

    previa = gerar_previa(2026, 3, db_com_dados)
    xlsx_bytes = gerar_excel(previa, 2026, 3)

    # Verifica que a palavra "alerta" não aparece em nenhuma célula do Excel
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                assert "alerta" not in cell.lower(), f"Célula com 'alerta' encontrada: {cell}"


def test_excel_34_linhas_de_dados(db_com_dados):
    previa = gerar_previa(2026, 3, db_com_dados)
    xlsx_bytes = gerar_excel(previa, 2026, 3)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Conta linhas de dados (a partir da linha 3, até a linha de total)
    linhas_dados = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        nome = row[1]  # coluna B (índice 1 no tuple)
        if nome and isinstance(nome, str) and len(nome.strip()) > 2:
            linhas_dados += 1

    assert linhas_dados == 34


# ---------------------------------------------------------------------------
# Testes do controle de período
# ---------------------------------------------------------------------------

def test_periodo_criado_como_aberto(db):
    from backend.routers.reports import _get_or_create_period
    period = _get_or_create_period(2026, 3, db)
    assert period.status == StatusPeriodo.aberto


def test_periodo_fechado_bloqueia_alteracoes(db):
    from backend.routers.reports import _get_or_create_period, _check_periodo_aberto
    from fastapi import HTTPException

    period = _get_or_create_period(2026, 3, db)
    period.status = StatusPeriodo.fechado
    db.commit()

    period_recarregado = _get_or_create_period(2026, 3, db)
    with pytest.raises(HTTPException) as exc:
        _check_periodo_aberto(period_recarregado)
    assert exc.value.status_code == 423
